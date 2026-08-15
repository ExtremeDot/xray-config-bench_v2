# -*- coding: utf-8 -*-
"""
CF Xray IP Benchmark
====================
Swap Cloudflare IPs into Xray share-links and rank them.

Usage:
  python cf_xray_benchmark.py
  python cf_xray_benchmark.py workers=1 max_ips=5 clear_temp=false
  python cf_xray_benchmark.py custom

Deps:
  pip install requests[socks] openpyxl colorama
"""

from __future__ import annotations

import os
import sys
import json
import time
import base64
import socket
import ipaddress
import statistics
import subprocess
import zipfile
import platform
import urllib.parse
import concurrent.futures
import threading
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
from urllib.request import urlretrieve

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    from colorama import init as colorama_init, Fore, Style
    colorama_init(autoreset=True)
except ImportError:
    class _D:
        def __getattr__(self, _):
            return ""
    Fore = Style = _D()  # type: ignore


# ==================== Settings ====================

SCRIPT_DIR = Path(__file__).resolve().parent
TEMP_DIR = SCRIPT_DIR / "temp"
RESULTS_DIR = SCRIPT_DIR / "results"
LOG_DIR = SCRIPT_DIR / "logs"

DEFAULT_WORKERS = 1
LATENCY_SAMPLES = 5
DOWNLOAD_BYTES = 250_000
DOWNLOAD_ROUNDS = 2
UPLOAD_BYTES = 80_000
UPLOAD_ROUNDS = 1
HTTP_TIMEOUT = 12
XRAY_WAIT = 3.0

RELAY_SITES = [
    ("Google", "https://www.google.com/generate_204"),
    ("Cloudflare", "https://www.cloudflare.com/cdn-cgi/trace"),
    ("YouTube", "https://www.youtube.com/generate_204"),
    ("Instagram", "https://www.instagram.com/"),
    ("GitHub", "https://github.com/"),
    ("Microsoft", "https://www.microsoft.com/"),
]

XRAY_WIN_URL = "https://github.com/XTLS/Xray-core/releases/download/v26.3.27/Xray-windows-64.zip"
XRAY_LIN_URL = "https://github.com/XTLS/Xray-core/releases/download/v26.3.27/Xray-linux-64.zip"
GEOIP_URL = "https://github.com/Loyalsoldier/v2ray-rules-dat/releases/latest/download/geoip.dat"
GEOSITE_URL = "https://github.com/Loyalsoldier/v2ray-rules-dat/releases/latest/download/geosite.dat"

_log_lock = threading.Lock()
_log_path: Optional[Path] = None
_print_lock = threading.Lock()
CFG: Dict[str, Any] = {}


def load_config() -> Dict[str, Any]:
    """Load config.json next to script. Missing keys fall back to defaults."""
    defaults = {
        "workers": 1,
        "max_ips": 0,
        "report_name": "",
        "clear_temp": None,
        "display": {
            "show_progress": True,
            "show_live_result": True,
            "show_final_tables": True,
            "colored_output": True,
            "write_log_file": True,
        },
        "baseline": {"enabled": True},
        "tests": {
            "latency": {"enabled": True, "samples": 5},
            "download": {"enabled": True, "bytes": 250000, "rounds": 2},
            "upload": {"enabled": True, "bytes": 80000, "rounds": 1},
        },
        "relay": {
            "enabled": True,
            "samples_per_site": 2,
            "sites": {
                "Google": {"enabled": True, "url": "https://www.google.com/generate_204"},
                "Cloudflare": {"enabled": True, "url": "https://www.cloudflare.com/cdn-cgi/trace"},
                "YouTube": {"enabled": True, "url": "https://www.youtube.com/generate_204"},
                "Instagram": {"enabled": True, "url": "https://www.instagram.com/"},
                "GitHub": {"enabled": True, "url": "https://github.com/"},
                "Microsoft": {"enabled": True, "url": "https://www.microsoft.com/"},
            },
        },
        "scoring": {
            "web": True,
            "instagram": True,
            "gaming": True,
            "overall_weights": {"web": 0.35, "instagram": 0.35, "gaming": 0.30},
        },
        "ranking": {
            "top_n": 5,
            "sort_by": "overall",
            "show_web_top": True,
            "show_instagram_top": True,
            "show_gaming_top": True,
            "show_overall_top": True,
            "compare_with_baseline": True,
        },
        "timeouts": {"http_seconds": 12, "xray_startup_seconds": 3.0},
        "paths": {
            "cfip_file": "cfip.txt",
            "links_file": "links.txt",
            "temp_dir": "temp",
            "results_dir": "results",
            "logs_dir": "logs",
        },
        "excel": {"enabled": True},
        "filter": {
            "enabled": True,
            "max_latency_ms": 2000,
            "quick_timeout_seconds": 3,
            "require_exit_ip": False,
        },
    }

    path = SCRIPT_DIR / "config.json"
    if not path.exists():
        warn("config.json not found — using built-in defaults")
        return defaults

    try:
        user = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        warn(f"config.json invalid ({e}) — using defaults")
        return defaults

    def merge(base, override):
        out = dict(base)
        for k, v in override.items():
            if k.startswith("_"):
                continue
            if isinstance(v, dict) and isinstance(out.get(k), dict):
                out[k] = merge(out[k], v)
            else:
                out[k] = v
        return out

    cfg = merge(defaults, user)
    ok(f"Loaded config: {path.name}")
    return cfg


def cfg_get(*keys, default=None):
    cur = CFG
    for k in keys:
        if not isinstance(cur, dict) or k not in cur:
            return default
        cur = cur[k]
    return cur


# ==================== Colors ====================

def c(text: str, color: str = "") -> str:
    return f"{color}{text}{Style.RESET_ALL}"


def info(msg: str) -> None:
    print(c("[INFO] ", Fore.CYAN) + msg)


def ok(msg: str) -> None:
    print(c("[ OK ] ", Fore.GREEN) + msg)


def warn(msg: str) -> None:
    print(c("[WARN] ", Fore.YELLOW) + msg)


def err(msg: str) -> None:
    print(c("[ERR ] ", Fore.RED) + msg)


def progress(msg: str) -> None:
    """Lightweight live progress (does not look frozen)."""
    if not cfg_get("display", "show_progress", default=True):
        return
    print(c("  ... ", Fore.BLUE) + msg, flush=True)


def title(msg: str) -> None:
    print()
    print(c("=" * 64, Fore.MAGENTA))
    print(c(f"  {msg}", Fore.MAGENTA + Style.BRIGHT))
    print(c("=" * 64, Fore.MAGENTA))


# ==================== Log ====================

def init_log(report: str = "") -> Path:
    global _log_path
    LOG_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    prefix = f"{report}_" if report else ""
    _log_path = LOG_DIR / f"{prefix}benchmark_{ts}.log"
    with open(_log_path, "w", encoding="utf-8") as f:
        f.write(f"CF Xray IP Benchmark\nStarted: {datetime.now().isoformat()}\n")
        f.write("=" * 72 + "\n\n")
    info(f"Log: {_log_path}")
    return _log_path


def log(text: str) -> None:
    if not _log_path or not cfg_get("display", "write_log_file", default=True):
        return
    with _log_lock:
        try:
            with open(_log_path, "a", encoding="utf-8") as f:
                f.write(text + "\n")
                f.flush()
        except Exception:
            pass


# ==================== Helpers ====================

def is_windows() -> bool:
    return platform.system().lower() == "windows"


def xray_name() -> str:
    return "xray.exe" if is_windows() else "xray"


def creation_flags() -> int:
    if is_windows():
        return getattr(subprocess, "CREATE_NO_WINDOW", 0)
    return 0


def free_port() -> int:
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.bind(("127.0.0.1", 0))
    port = s.getsockname()[1]
    s.close()
    return port


def port_open(port: int, timeout: float = 0.4) -> bool:
    try:
        with socket.create_connection(("127.0.0.1", port), timeout=timeout):
            return True
    except OSError:
        return False


def download(url: str, dest: Path, label: str) -> bool:
    warn(f"{label} missing, downloading...")
    try:
        def hook(count, block, total):
            if total > 0:
                pct = min(100, count * block * 100 // total)
                print(f"\r  {label}: {pct}%", end="", flush=True)
        urlretrieve(url, dest, reporthook=hook)
        print()
        ok(f"{label} ready")
        return True
    except Exception as e:
        err(f"Download failed ({label}): {e}")
        return False


def ensure_xray() -> str:
    path = SCRIPT_DIR / xray_name()
    if path.exists():
        ok(f"Found {path.name}")
        return str(path)

    url = XRAY_WIN_URL if is_windows() else XRAY_LIN_URL
    zpath = SCRIPT_DIR / "xray_download.zip"
    warn("xray not found, downloading...")
    try:
        urlretrieve(url, zpath)
        with zipfile.ZipFile(zpath, "r") as zf:
            for member in zf.namelist():
                low = member.lower()
                if low.endswith("xray.exe") or low.endswith("/xray") or low == "xray":
                    zf.extract(member, path=SCRIPT_DIR)
                    extracted = SCRIPT_DIR / member
                    if extracted.resolve() != path.resolve():
                        if path.exists():
                            path.unlink()
                        extracted.replace(path)
                    break
            else:
                zf.extractall(SCRIPT_DIR)
                for cand in SCRIPT_DIR.rglob("*"):
                    if cand.is_file() and cand.name.lower() in ("xray.exe", "xray"):
                        if cand.resolve() != path.resolve():
                            cand.replace(path)
                        break
        if zpath.exists():
            zpath.unlink(missing_ok=True)
        if not path.exists():
            raise FileNotFoundError("xray binary missing after extract")
        if not is_windows():
            os.chmod(path, 0o755)
        ok("xray ready")
        return str(path)
    except Exception as e:
        err(f"Cannot get xray: {e}")
        sys.exit(1)


def ensure_geodata() -> None:
    for name, url in (("geoip.dat", GEOIP_URL), ("geosite.dat", GEOSITE_URL)):
        dest = SCRIPT_DIR / name
        if dest.exists():
            ok(f"{name} found")
        else:
            download(url, dest, name)


# ==================== Loaders ====================

def load_cf_ips(path: Path, max_ips: Optional[int]) -> List[str]:
    """
    Supported lines in cfip.txt:
      1181 - 0 - 172.66.170.97     (scan format: dl - ul - IP)
      172.66.170.97               (plain IPv4)
      172.66.170.0/24             (CIDR — expanded)
      domain.example.com          (domain — resolved via DNS)
      # comment
    """
    if not path.exists():
        err(f"Missing {path.name}")
        return []

    def is_ipv4(s: str) -> bool:
        try:
            ipaddress.IPv4Address(s)
            return True
        except Exception:
            return False

    def expand_cidr(cidr: str) -> List[str]:
        try:
            net = ipaddress.ip_network(cidr, strict=False)
            # Safety: very large nets would explode memory/time
            if net.num_addresses > 4096:
                warn(f"CIDR {cidr} has {net.num_addresses} addresses — sampling first 4096")
                out = []
                for i, host in enumerate(net.hosts()):
                    if i >= 4096:
                        break
                    out.append(str(host))
                return out
            # .hosts() skips network/broadcast; for /32 use the single address
            if net.num_addresses == 1:
                return [str(net.network_address)]
            return [str(h) for h in net.hosts()]
        except Exception as e:
            warn(f"Bad CIDR '{cidr}': {e}")
            return []

    def resolve_domain(host: str) -> List[str]:
        try:
            infos = socket.getaddrinfo(host, None, socket.AF_INET)
            found = []
            for info in infos:
                ip = info[4][0]
                if is_ipv4(ip) and ip not in found:
                    found.append(ip)
            if found:
                ok(f"Resolved {host} -> {', '.join(found)}")
            else:
                warn(f"No A record for {host}")
            return found
        except Exception as e:
            warn(f"DNS failed for {host}: {e}")
            return []

    ips: List[str] = []
    for raw in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue

        # Format: "dl - ul - IP"  (dashes may be en-dash)
        token = line
        if " - " in line.replace("\u2013", "-").replace("–", " - "):
            parts = [x.strip() for x in line.replace("–", "-").split("-")]
            if len(parts) >= 3:
                token = parts[-1].strip()
        else:
            # single token (maybe with trailing comment)
            token = line.split()[0].strip()

        token = token.strip().strip(",")

        if not token:
            continue

        # CIDR
        if "/" in token and not token.startswith("http"):
            expanded = expand_cidr(token)
            ips.extend(expanded)
            if expanded:
                info(f"CIDR {token} -> {len(expanded)} IP(s)")
            continue

        # Plain IPv4
        if is_ipv4(token):
            ips.append(token)
            continue

        # Domain (basic check: has a letter and a dot)
        if any(c.isalpha() for c in token) and "." in token:
            ips.extend(resolve_domain(token))
            continue

        warn(f"Skipped unknown entry: {token}")

    seen = set()
    uniq: List[str] = []
    for ip in ips:
        if ip not in seen:
            seen.add(ip)
            uniq.append(ip)

    if max_ips and len(uniq) > max_ips:
        uniq = uniq[:max_ips]
        warn(f"Using first {max_ips} IPs only")
    ok(f"Loaded {len(uniq)} targets from {path.name}")
    return uniq


def load_links(path: Path) -> List[str]:
    if not path.exists():
        err(f"Missing {path.name}")
        return []
    links = []
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "://" in line:
            links.append(line)
    ok(f"Loaded {len(links)} links")
    return links


# ==================== Parsers ====================

def parse_vless(link: str) -> Optional[Dict]:
    try:
        rest = link[8:]
        remark = "vless"
        if "#" in rest:
            rest, remark = rest.rsplit("#", 1)
            remark = urllib.parse.unquote(remark)
        if "?" in rest:
            main, query = rest.split("?", 1)
            params = dict(urllib.parse.parse_qsl(query))
        else:
            main, params = rest, {}
        uuid, server_part = main.split("@", 1)
        address, port_s = server_part.rsplit(":", 1)
        port = int(port_s)
        network = params.get("type", "tcp")
        security = params.get("security", "none")
        sni = params.get("sni") or params.get("host") or address
        fp = params.get("fp", "")
        alpn = params.get("alpn", "")
        flow = params.get("flow", "")
        path = params.get("path", "")
        host = params.get("host", "")
        service_name = params.get("serviceName", "")
        insecure = params.get("allowInsecure", params.get("insecure", "0")) == "1"

        stream: Dict[str, Any] = {"network": network, "security": security}
        if security in ("tls", "reality"):
            tls: Dict[str, Any] = {"serverName": sni, "allowInsecure": insecure}
            if fp:
                tls["fingerprint"] = fp
            if alpn:
                tls["alpn"] = alpn.split(",")
            if security == "reality":
                tls["publicKey"] = params.get("pbk", "")
                tls["shortId"] = params.get("sid", "")
                tls["spiderX"] = params.get("spx", "")
                stream["realitySettings"] = tls
            else:
                stream["tlsSettings"] = tls
        if network == "ws":
            stream["wsSettings"] = {"path": path or "/", "headers": {"Host": host or sni}}
        elif network == "grpc":
            stream["grpcSettings"] = {
                "serviceName": service_name,
                "multiMode": params.get("mode", "gun") == "multi",
            }
        elif network == "httpupgrade":
            stream["httpupgradeSettings"] = {"path": path or "/", "host": host or sni}
        elif network == "xhttp":
            xh: Dict[str, Any] = {"path": path or "/", "host": host or sni}
            if params.get("mode"):
                xh["mode"] = params["mode"]
            stream["xhttpSettings"] = xh

        outbound = {
            "tag": "proxy",
            "protocol": "vless",
            "settings": {
                "vnext": [{
                    "address": address,
                    "port": port,
                    "users": [{
                        "id": uuid,
                        "encryption": params.get("encryption", "none"),
                        "flow": flow,
                    }],
                }]
            },
            "streamSettings": stream,
        }
        return {"remark": remark, "outbound": outbound, "address": address, "port": port}
    except Exception as e:
        warn(f"VLESS parse error: {e}")
        return None


def parse_trojan(link: str) -> Optional[Dict]:
    try:
        rest = link[9:]
        remark = "trojan"
        if "#" in rest:
            rest, remark = rest.rsplit("#", 1)
            remark = urllib.parse.unquote(remark)
        if "?" in rest:
            main, query = rest.split("?", 1)
            params = dict(urllib.parse.parse_qsl(query))
        else:
            main, params = rest, {}
        password, server_part = main.split("@", 1)
        address, port_s = server_part.rsplit(":", 1)
        port = int(port_s)
        sni = params.get("sni") or params.get("peer") or address
        network = params.get("type", "tcp")
        fp = params.get("fp", "")
        path = params.get("path", "")
        host = params.get("host", "")
        insecure = params.get("allowInsecure", "0") == "1"

        stream: Dict[str, Any] = {
            "network": network,
            "security": "tls",
            "tlsSettings": {"serverName": sni, "allowInsecure": insecure},
        }
        if fp:
            stream["tlsSettings"]["fingerprint"] = fp
        if network == "ws":
            stream["wsSettings"] = {"path": path or "/", "headers": {"Host": host or sni}}

        outbound = {
            "tag": "proxy",
            "protocol": "trojan",
            "settings": {"servers": [{"address": address, "port": port, "password": password}]},
            "streamSettings": stream,
        }
        return {"remark": remark, "outbound": outbound, "address": address, "port": port}
    except Exception as e:
        warn(f"Trojan parse error: {e}")
        return None


def parse_vmess(link: str) -> Optional[Dict]:
    try:
        b64 = link[8:] + "=" * (-len(link[8:]) % 4)
        data = json.loads(base64.b64decode(b64).decode("utf-8"))
        address = data.get("add")
        port = int(data.get("port", 443))
        uuid = data.get("id")
        network = data.get("net", "tcp")
        tls = data.get("tls", "")
        sni = data.get("sni") or data.get("host") or address
        path = data.get("path", "")
        host = data.get("host", "")
        remark = data.get("ps", "vmess")

        stream: Dict[str, Any] = {"network": network}
        if tls in ("tls", "reality"):
            stream["security"] = tls
            key = "tlsSettings" if tls == "tls" else "realitySettings"
            stream[key] = {"serverName": sni, "allowInsecure": False}
        else:
            stream["security"] = "none"
        if network == "ws":
            stream["wsSettings"] = {"path": path or "/", "headers": {"Host": host or sni}}

        outbound = {
            "tag": "proxy",
            "protocol": "vmess",
            "settings": {
                "vnext": [{
                    "address": address,
                    "port": port,
                    "users": [{
                        "id": uuid,
                        "alterId": int(data.get("aid", 0)),
                        "security": data.get("scy", "auto"),
                    }],
                }]
            },
            "streamSettings": stream,
        }
        return {"remark": remark, "outbound": outbound, "address": address, "port": port}
    except Exception as e:
        warn(f"VMess parse error: {e}")
        return None


def parse_link(link: str) -> Optional[Dict]:
    link = link.strip()
    if link.startswith("vless://"):
        return parse_vless(link)
    if link.startswith("trojan://"):
        return parse_trojan(link)
    if link.startswith("vmess://"):
        return parse_vmess(link)
    warn(f"Unsupported link: {link[:40]}...")
    return None


def build_config(parsed: Dict, cf_ip: str, socks_port: int) -> Dict:
    outbound = json.loads(json.dumps(parsed["outbound"]))
    if outbound["protocol"] in ("vless", "vmess"):
        outbound["settings"]["vnext"][0]["address"] = cf_ip
    elif outbound["protocol"] == "trojan":
        outbound["settings"]["servers"][0]["address"] = cf_ip

    return {
        "log": {"loglevel": "warning"},
        "inbounds": [{
            "tag": "socks",
            "port": socks_port,
            "listen": "127.0.0.1",
            "protocol": "mixed",
            "settings": {"auth": "noauth", "udp": True},
        }],
        "outbounds": [
            outbound,
            {"tag": "direct", "protocol": "freedom"},
        ],
        "routing": {
            "domainStrategy": "AsIs",
            "rules": [{"type": "field", "outboundTag": "proxy", "network": "tcp,udp"}],
        },
    }


# ==================== HTTP tests ====================

def req(method: str, url: str, proxies: Dict, **kw) -> Optional[requests.Response]:
    try:
        to = kw.pop("timeout", None)
        if to is None:
            to = float(cfg_get("timeouts", "http_seconds", default=12) or 12)
        return requests.request(method, url, proxies=proxies, timeout=to, **kw)
    except Exception:
        return None


def quick_probe(proxies: Dict) -> Dict[str, Any]:
    """
    Fast connectivity check. Returns:
      ok, latency_ms, exit_ip, loc, colo, reason
    """
    qto = float(cfg_get("filter", "quick_timeout_seconds", default=3) or 3)
    max_ms = float(cfg_get("filter", "max_latency_ms", default=2000) or 2000)
    out: Dict[str, Any] = {
        "ok": False, "latency_ms": None,
        "exit_ip": "?", "loc": "?", "colo": "?", "reason": "",
    }

    t0 = time.perf_counter()
    r = req("GET", "https://www.cloudflare.com/cdn-cgi/trace", proxies, timeout=qto)
    ms = (time.perf_counter() - t0) * 1000
    out["latency_ms"] = round(ms, 1)

    if r is None or r.status_code != 200:
        # fallback ipify once
        t1 = time.perf_counter()
        r2 = req("GET", "https://api.ipify.org?format=json", proxies, timeout=qto)
        ms2 = (time.perf_counter() - t1) * 1000
        out["latency_ms"] = round(ms2, 1)
        if r2 is None or r2.status_code != 200:
            out["reason"] = f"no response within {qto}s"
            return out
        try:
            out["exit_ip"] = r2.json().get("ip", "?")
        except Exception:
            out["exit_ip"] = "?"
    else:
        data = {}
        for line in r.text.strip().splitlines():
            if "=" in line:
                k, v = line.split("=", 1)
                data[k] = v
        out["exit_ip"] = data.get("ip", "?")
        out["loc"] = data.get("loc", "?")
        out["colo"] = data.get("colo", "?")

    if out["latency_ms"] is not None and out["latency_ms"] > max_ms:
        out["reason"] = f"ping {out['latency_ms']}ms > max {max_ms}ms"
        return out

    if cfg_get("filter", "require_exit_ip", default=False) and out["exit_ip"] in ("?", "", None):
        out["reason"] = "exit IP unknown"
        return out

    out["ok"] = True
    return out


def test_latency(proxies: Dict) -> Optional[Dict]:
    if not cfg_get("tests", "latency", "enabled", default=True):
        return None
    samples = int(cfg_get("tests", "latency", "samples", default=5) or 5)
    times = []
    for _ in range(samples):
        t0 = time.perf_counter()
        r = req("GET", "https://www.cloudflare.com/cdn-cgi/trace", proxies)
        ms = (time.perf_counter() - t0) * 1000
        if r is not None and r.status_code == 200:
            times.append(ms)
        time.sleep(0.08)
    if not times:
        return None
    return {
        "avg": round(statistics.mean(times), 1),
        "min": round(min(times), 1),
        "max": round(max(times), 1),
        "jitter": round(statistics.stdev(times), 1) if len(times) > 1 else 0.0,
        "loss": round((1 - len(times) / samples) * 100, 1),
    }


def test_download(proxies: Dict) -> Optional[Dict]:
    if not cfg_get("tests", "download", "enabled", default=True):
        return None
    nbytes = int(cfg_get("tests", "download", "bytes", default=250000) or 250000)
    rounds = int(cfg_get("tests", "download", "rounds", default=2) or 2)
    speeds = []
    for _ in range(rounds):
        t0 = time.perf_counter()
        r = req("GET", f"https://speed.cloudflare.com/__down?bytes={nbytes}", proxies)
        elapsed = time.perf_counter() - t0
        if r is not None and r.status_code == 200 and elapsed > 0.05:
            speeds.append((nbytes * 8) / (elapsed * 1_000_000))
        time.sleep(0.1)
    if not speeds:
        return None
    return {
        "avg": round(statistics.mean(speeds), 2),
        "min": round(min(speeds), 2),
        "max": round(max(speeds), 2),
    }


def test_upload(proxies: Dict) -> Optional[Dict]:
    if not cfg_get("tests", "upload", "enabled", default=True):
        return None
    nbytes = int(cfg_get("tests", "upload", "bytes", default=80000) or 80000)
    rounds = int(cfg_get("tests", "upload", "rounds", default=1) or 1)
    data = os.urandom(nbytes)
    speeds = []
    for _ in range(rounds):
        t0 = time.perf_counter()
        r = req(
            "POST",
            "https://speed.cloudflare.com/__up",
            proxies,
            data=data,
            headers={"Content-Type": "application/octet-stream"},
        )
        elapsed = time.perf_counter() - t0
        if r is not None and r.status_code in (200, 204) and elapsed > 0.05:
            speeds.append((nbytes * 8) / (elapsed * 1_000_000))
        time.sleep(0.1)
    if not speeds:
        return None
    return {
        "avg": round(statistics.mean(speeds), 2),
        "min": round(min(speeds), 2),
        "max": round(max(speeds), 2),
    }


def exit_info(proxies: Dict) -> Dict:
    r = req("GET", "https://www.cloudflare.com/cdn-cgi/trace", proxies)
    if r is not None and r.status_code == 200:
        data = {}
        for line in r.text.strip().splitlines():
            if "=" in line:
                k, v = line.split("=", 1)
                data[k] = v
        return {"ip": data.get("ip", "?"), "loc": data.get("loc", "?"), "colo": data.get("colo", "?")}
    r = req("GET", "https://api.ipify.org?format=json", proxies)
    if r is not None and r.status_code == 200:
        try:
            return {"ip": r.json().get("ip", "?"), "loc": "?", "colo": "?"}
        except Exception:
            pass
    return {"ip": "?", "loc": "?", "colo": "?"}


def test_relay(proxies: Dict) -> Dict[str, Optional[float]]:
    out: Dict[str, Optional[float]] = {}
    if not cfg_get("relay", "enabled", default=True):
        return out
    sites = cfg_get("relay", "sites", default={}) or {}
    n_samples = int(cfg_get("relay", "samples_per_site", default=2) or 2)
    for name, meta in sites.items():
        if isinstance(meta, dict):
            if not meta.get("enabled", True):
                continue
            url = meta.get("url")
        else:
            url = meta
        if not url:
            continue
        samples = []
        for _ in range(n_samples):
            t0 = time.perf_counter()
            r = req("GET", url, proxies, allow_redirects=True)
            ms = (time.perf_counter() - t0) * 1000
            if r is not None and r.status_code < 500:
                samples.append(ms)
            time.sleep(0.05)
        out[name] = round(statistics.mean(samples), 1) if samples else None
    return out


# ==================== Scoring ====================

def score_web(lat, jitter, loss, dl, ul) -> float:
    if lat is None or dl is None:
        return 0.0
    s = 10.0
    if lat > 300:
        s -= 4
    elif lat > 200:
        s -= 2.5
    elif lat > 120:
        s -= 1.5
    elif lat > 80:
        s -= 0.8
    if jitter > 50:
        s -= 2
    elif jitter > 30:
        s -= 1
    if loss > 8:
        s -= 2.5
    elif loss > 3:
        s -= 1
    if dl < 1:
        s -= 3
    elif dl < 3:
        s -= 1.5
    elif dl < 8:
        s -= 0.5
    if ul is not None and ul < 0.5:
        s -= 1
    return max(0.0, round(s, 1))


def score_instagram(lat, jitter, loss, dl, ul) -> float:
    if lat is None or dl is None:
        return 0.0
    s = 10.0
    if lat > 250:
        s -= 4
    elif lat > 150:
        s -= 2.5
    elif lat > 100:
        s -= 1.5
    elif lat > 70:
        s -= 0.8
    if jitter > 40:
        s -= 2.5
    elif jitter > 25:
        s -= 1.2
    if loss > 5:
        s -= 3
    elif loss > 2:
        s -= 1.5
    if dl < 2:
        s -= 2.5
    elif dl < 5:
        s -= 1
    if ul is None or ul < 0.8:
        s -= 2
    elif ul < 2:
        s -= 0.8
    return max(0.0, round(s, 1))


def score_gaming(lat, jitter, loss, dl, ul) -> float:
    """Gaming prioritizes low latency + low jitter + low loss (score 0-10)."""
    if lat is None:
        return 0.0
    s = 10.0
    if lat > 180:
        s -= 5
    elif lat > 120:
        s -= 3.5
    elif lat > 80:
        s -= 2
    elif lat > 50:
        s -= 1
    if jitter > 40:
        s -= 3
    elif jitter > 25:
        s -= 2
    elif jitter > 15:
        s -= 1
    if loss > 5:
        s -= 3
    elif loss > 2:
        s -= 1.5
    elif loss > 0.5:
        s -= 0.5
    # bandwidth less critical for gaming, mild penalty only
    if dl is not None and dl < 1:
        s -= 1
    return max(0.0, round(s, 1))


def score_overall(web, insta, gaming) -> float:
    w = cfg_get("scoring", "overall_weights", default={}) or {}
    ww = float(w.get("web", 0.35))
    wi = float(w.get("instagram", 0.35))
    wg = float(w.get("gaming", 0.30))
    # if a score type disabled, redistribute
    if not cfg_get("scoring", "web", default=True):
        ww = 0
    if not cfg_get("scoring", "instagram", default=True):
        wi = 0
    if not cfg_get("scoring", "gaming", default=True):
        wg = 0
    total = ww + wi + wg
    if total <= 0:
        return 0.0
    return round((web * ww + insta * wi + gaming * wg) / total, 1)


# ==================== Baseline (direct, no proxy) ====================

def measure_baseline() -> Dict[str, Any]:
    """Test user's own internet quality without any proxy."""
    empty = {
        "lat_avg": None, "jitter": None, "loss": None,
        "dl_avg": None, "ul_avg": None,
        "web": 0, "insta": 0, "gaming": 0, "overall": 0, "ok": False,
    }
    if not cfg_get("baseline", "enabled", default=True):
        info("Baseline test disabled in config.json")
        return empty
    title("Baseline: your direct internet (no proxy)")
    info("Measuring your real connection quality before proxy tests...")
    proxies = {}  # direct
    base: Dict[str, Any] = {
        "lat_avg": None, "jitter": None, "loss": None,
        "dl_avg": None, "ul_avg": None,
        "web": 0, "insta": 0, "gaming": 0, "overall": 0,
        "ok": False,
    }

    progress("baseline | latency...")
    lat = test_latency(proxies)
    progress("baseline | download...")
    dl = test_download(proxies)
    progress("baseline | upload...")
    ul = test_upload(proxies)

    if lat:
        base["lat_avg"] = lat["avg"]
        base["jitter"] = lat["jitter"]
        base["loss"] = lat["loss"]
    if dl:
        base["dl_avg"] = dl["avg"]
    if ul:
        base["ul_avg"] = ul["avg"]

    web = score_web(
        base["lat_avg"], base["jitter"] or 0, base["loss"] or 0,
        base["dl_avg"], base["ul_avg"],
    )
    insta = score_instagram(
        base["lat_avg"], base["jitter"] or 0, base["loss"] or 0,
        base["dl_avg"], base["ul_avg"],
    )
    gaming = score_gaming(
        base["lat_avg"], base["jitter"] or 0, base["loss"] or 0,
        base["dl_avg"], base["ul_avg"],
    )
    base["web"] = web
    base["insta"] = insta
    base["gaming"] = gaming
    base["overall"] = score_overall(web, insta, gaming)
    base["ok"] = lat is not None or dl is not None

    print()
    if base["ok"]:
        ok("Your direct connection:")
        print(f"  Latency   : {c(str(base['lat_avg'] or 'N/A'), Fore.GREEN)} ms  "
              f"jitter={base['jitter']}  loss={base['loss']}%")
        print(f"  Download  : {c(str(base['dl_avg'] or 'N/A'), Fore.GREEN)} Mbps")
        print(f"  Upload    : {c(str(base['ul_avg'] or 'N/A'), Fore.GREEN)} Mbps")
        print(f"  Scores    : Web={base['web']}/10  Instagram={base['insta']}/10  "
              f"Gaming={base['gaming']}/10  Overall={c(str(base['overall'])+'/10', Fore.YELLOW + Style.BRIGHT)}")
        info("Scores are out of 10 (higher = better). This is your ceiling without proxy overhead.")
    else:
        warn("Could not measure baseline (network issue?). Continuing without comparison.")
    print(c("-" * 64, Fore.BLUE))
    log(f"BASELINE lat={base['lat_avg']} dl={base['dl_avg']} ul={base['ul_avg']} "
        f"web={base['web']} insta={base['insta']} gaming={base['gaming']}")
    return base


# ==================== One test ====================

def run_one(xray_path: str, parsed: Dict, cf_ip: str, remark: str) -> Dict:
    socks_port = free_port()
    proxies = {
        "http": f"socks5h://127.0.0.1:{socks_port}",
        "https": f"socks5h://127.0.0.1:{socks_port}",
    }
    result: Dict[str, Any] = {
        "name": f"{remark} | {cf_ip}",
        "cf_ip": cf_ip,
        "remark": remark,
        "status": "FAIL",
        "exit_ip": "?",
        "loc": "?",
        "colo": "?",
        "lat_avg": None,
        "lat_min": None,
        "lat_max": None,
        "jitter": None,
        "loss": None,
        "dl_avg": None,
        "dl_min": None,
        "dl_max": None,
        "ul_avg": None,
        "ul_min": None,
        "ul_max": None,
        "web_score": 0,
        "insta_score": 0,
        "gaming_score": 0,
        "overall": 0,
        "relay": {},
        "port": socks_port,
    }

    TEMP_DIR.mkdir(exist_ok=True)
    conf_path = TEMP_DIR / f"cfg_{cf_ip.replace('.', '_')}_{socks_port}.json"
    err_path = TEMP_DIR / f"cfg_{cf_ip.replace('.', '_')}_{socks_port}.err"
    proc = None

    try:
        config = build_config(parsed, cf_ip, socks_port)
        conf_path.write_text(json.dumps(config, indent=2), encoding="utf-8")
        progress(f"{cf_ip} | starting xray on port {socks_port}...")

        err_fh = open(err_path, "w", encoding="utf-8")
        proc = subprocess.Popen(
            [xray_path, "run", "-c", str(conf_path)],
            stdout=err_fh,
            stderr=err_fh,
            cwd=str(SCRIPT_DIR),
            creationflags=creation_flags(),
        )
        err_fh.close()

        ready = False
        loops = int(float(cfg_get("timeouts", "xray_startup_seconds", default=3.0) or 3.0) / 0.25) + 4
        for _ in range(loops):
            time.sleep(0.25)
            if proc.poll() is not None:
                break
            if port_open(socks_port):
                ready = True
                break

        if proc.poll() is not None or not ready:
            detail = ""
            if err_path.exists():
                detail = err_path.read_text(encoding="utf-8", errors="replace").strip()[-400:]
            if not detail:
                detail = f"alive={proc.poll() is None} port={socks_port} ready={ready}"
            result["status"] = f"FAIL (start: {detail})"
            log(f"FAIL start {cf_ip} port={socks_port}\n{detail}")
            try:
                failed = TEMP_DIR / f"FAILED_{cf_ip.replace('.', '_')}.json"
                failed.write_text(conf_path.read_text(encoding="utf-8"), encoding="utf-8")
            except Exception:
                pass
            return result

        # ---- quick filter: skip dead / ultra-slow IPs early ----
        if cfg_get("filter", "enabled", default=True):
            progress(f"{cf_ip} | quick probe (skip if dead or ping too high)...")
            probe = quick_probe(proxies)
            result["exit_ip"] = probe.get("exit_ip", "?")
            result["loc"] = probe.get("loc", "?")
            result["colo"] = probe.get("colo", "?")
            if not probe.get("ok"):
                reason = probe.get("reason") or "unreachable"
                result["status"] = f"SKIP ({reason})"
                if probe.get("latency_ms") is not None:
                    result["lat_avg"] = probe["latency_ms"]
                    result["lat_min"] = probe["latency_ms"]
                    result["lat_max"] = probe["latency_ms"]
                log(f"SKIP {cf_ip} {reason}")
                progress(f"{cf_ip} | skipped — {reason}")
                return result
            progress(f"{cf_ip} | probe ok ({probe.get('latency_ms')}ms) exit={probe.get('exit_ip')}")
        else:
            progress(f"{cf_ip} | checking exit IP...")
            ex = exit_info(proxies)
            result["exit_ip"] = ex["ip"]
            result["loc"] = ex["loc"]
            result["colo"] = ex["colo"]

        progress(f"{cf_ip} | measuring latency...")
        lat = test_latency(proxies)
        if lat:
            result["lat_avg"] = lat["avg"]
            result["lat_min"] = lat["min"]
            result["lat_max"] = lat["max"]
            result["jitter"] = lat["jitter"]
            result["loss"] = lat["loss"]
            # also skip after full latency if still above threshold
            max_ms = float(cfg_get("filter", "max_latency_ms", default=2000) or 2000)
            if cfg_get("filter", "enabled", default=True) and lat["avg"] > max_ms:
                result["status"] = f"SKIP (avg ping {lat['avg']}ms > {max_ms}ms)"
                log(f"SKIP {cf_ip} avg latency {lat['avg']}")
                progress(f"{cf_ip} | skipped — avg ping too high")
                return result
        elif cfg_get("tests", "latency", "enabled", default=True):
            result["status"] = "FAIL (latency)"
            log(f"FAIL latency {cf_ip}")
            return result

        progress(f"{cf_ip} | download / upload speed...")
        dl = test_download(proxies)
        ul = test_upload(proxies)
        if cfg_get("tests", "latency", "enabled", default=True) is False and not dl and not ul:
            result["status"] = "FAIL (no tests enabled or all failed)"
            return result
        if dl:
            result["dl_avg"] = dl["avg"]
            result["dl_min"] = dl["min"]
            result["dl_max"] = dl["max"]
        if ul:
            result["ul_avg"] = ul["avg"]
            result["ul_min"] = ul["min"]
            result["ul_max"] = ul["max"]

        if cfg_get("relay", "enabled", default=True):
            progress(f"{cf_ip} | relay pings...")
        result["relay"] = test_relay(proxies)

        progress(f"{cf_ip} | scoring...")
        lat_avg = lat["avg"] if lat else None
        jitter = lat["jitter"] if lat else 0
        loss = lat["loss"] if lat else 0
        dl_avg = dl["avg"] if dl else None
        ul_avg = ul["avg"] if ul else None
        web = score_web(lat_avg, jitter, loss, dl_avg, ul_avg) if cfg_get("scoring", "web", default=True) else 0.0
        insta = score_instagram(lat_avg, jitter, loss, dl_avg, ul_avg) if cfg_get("scoring", "instagram", default=True) else 0.0
        gaming = score_gaming(lat_avg, jitter, loss, dl_avg, ul_avg) if cfg_get("scoring", "gaming", default=True) else 0.0
        result["web_score"] = web
        result["insta_score"] = insta
        result["gaming_score"] = gaming
        result["overall"] = score_overall(web, insta, gaming)
        result["status"] = "OK"
        log(
            f"OK {cf_ip} lat={lat['avg']} dl={result['dl_avg']} ul={result['ul_avg']} "
            f"overall={result['overall']}"
        )

    except Exception as e:
        result["status"] = f"FAIL ({type(e).__name__}: {e})"
        log(f"EXCEPTION {cf_ip}: {e}")
    finally:
        if proc is not None and proc.poll() is None:
            try:
                proc.terminate()
                proc.wait(timeout=2)
            except Exception:
                try:
                    proc.kill()
                except Exception:
                    pass

    return result


def print_result(res: Dict, idx: int, total: int) -> None:
    if not cfg_get("display", "show_live_result", default=True):
        return
    ok_status = res["status"] == "OK"
    skip_status = str(res.get("status", "")).startswith("SKIP")
    if ok_status:
        st_col = Fore.GREEN + Style.BRIGHT
    elif skip_status:
        st_col = Fore.YELLOW + Style.BRIGHT
    else:
        st_col = Fore.RED + Style.BRIGHT
    with _print_lock:
        print()
        print(
            c(f"  [{idx}/{total}] ", Fore.WHITE + Style.BRIGHT)
            + c(res["cf_ip"], Fore.CYAN + Style.BRIGHT)
            + c(f"  |  {res['remark'][:36]}", Fore.WHITE)
        )
        print(c("  " + "-" * 60, Fore.BLUE))
        print(f"  Status   : {c(res['status'], st_col)}")
        print(
            f"  Exit IP  : {c(str(res['exit_ip']), Fore.YELLOW)}  "
            f"Loc={res['loc']}  Colo={res['colo']}  port={res.get('port', '?')}"
        )
        if res["lat_avg"] is not None:
            print(
                f"  Latency  : avg={c(str(res['lat_avg']), Fore.GREEN)} ms  "
                f"min={res['lat_min']} max={res['lat_max']}  "
                f"jitter={res['jitter']} loss={res['loss']}%"
            )
        if res["dl_avg"] is not None:
            print(f"  Download : avg={c(str(res['dl_avg']), Fore.GREEN)} Mbps")
        if res["ul_avg"] is not None:
            print(f"  Upload   : avg={c(str(res['ul_avg']), Fore.GREEN)} Mbps")
        print(
            f"  Scores   : Web={c(str(res['web_score']), Fore.CYAN)}/10  "
            f"Insta={c(str(res['insta_score']), Fore.MAGENTA)}/10  "
            f"Game={c(str(res.get('gaming_score', 0)), Fore.GREEN)}/10  "
            f"Overall={c(str(res['overall']), Fore.YELLOW + Style.BRIGHT)}/10"
        )
        relay = res.get("relay") or {}
        if relay:
            parts = []
            for k, v in relay.items():
                if v is not None:
                    parts.append(f"{k}={c(str(v) + 'ms', Fore.GREEN)}")
                else:
                    parts.append(f"{k}={c('fail', Fore.RED)}")
            print(f"  Relay    : {'  '.join(parts)}")
        print(c("  " + "-" * 60, Fore.BLUE))


# ==================== Excel ====================

def save_excel(results: List[Dict], report: str = "") -> Path:
    RESULTS_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    prefix = f"[{report}] " if report else ""
    path = RESULTS_DIR / f"{prefix}cf_ip_benchmark_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    headers = [
        "Rank", "Name", "CF IP", "Status", "Exit IP", "Loc", "Colo",
        "Lat Avg", "Lat Min", "Lat Max", "Jitter", "Loss %",
        "DL Avg", "DL Min", "DL Max", "UL Avg", "UL Min", "UL Max",
        "Web", "Instagram", "Gaming", "Overall",
        "Google", "Cloudflare", "YouTube", "InstagramR", "GitHub", "Microsoft",
    ]
    ws.append(headers)
    hf = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.fill = hf
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")

    ok_list = sorted(
        [r for r in results if r["status"] == "OK"],
        key=lambda x: x["overall"],
        reverse=True,
    )
    bad_list = [r for r in results if r["status"] != "OK"]
    ordered = ok_list + bad_list
    ok_fill = PatternFill("solid", fgColor="C6EFCE")
    bad_fill = PatternFill("solid", fgColor="FFC7CE")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for rank, r in enumerate(ordered, 1):
        relay = r.get("relay") or {}
        row = [
            rank if r["status"] == "OK" else "-",
            r["name"], r["cf_ip"], r["status"], r["exit_ip"], r["loc"], r["colo"],
            r["lat_avg"], r["lat_min"], r["lat_max"], r["jitter"], r["loss"],
            r["dl_avg"], r["dl_min"], r["dl_max"], r["ul_avg"], r["ul_min"], r["ul_max"],
            r["web_score"], r["insta_score"], r.get("gaming_score", 0), r["overall"],
            relay.get("Google"), relay.get("Cloudflare"), relay.get("YouTube"),
            relay.get("Instagram"), relay.get("GitHub"), relay.get("Microsoft"),
        ]
        ws.append(row)
        fill = ok_fill if r["status"] == "OK" else bad_fill
        for col in range(1, len(headers) + 1):
            cell = ws.cell(ws.max_row, col)
            cell.fill = fill
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.column_dimensions["B"].width = 34

    ws2 = wb.create_sheet("Ranking")
    ws2.append(["Rank", "CF IP", "Remark", "Overall", "Web", "Instagram", "Gaming", "Lat", "DL", "UL", "Exit IP"])
    for col in range(1, 12):
        ws2.cell(1, col).fill = hf
        ws2.cell(1, col).font = hfont
    for i, r in enumerate(ok_list, 1):
        ws2.append([
            i, r["cf_ip"], r["remark"], r["overall"], r["web_score"], r["insta_score"],
            r.get("gaming_score", 0),
            r["lat_avg"], r["dl_avg"], r["ul_avg"], r["exit_ip"],
        ])

    wb.save(path)
    ok(f"Excel: {path}")
    return path


# ==================== CLI ====================

def parse_args() -> Dict[str, Any]:
    args: Dict[str, Any] = {
        "workers": DEFAULT_WORKERS,
        "max_ips": None,
        "report": "",
        "custom": False,
        "clear_temp": None,
    }
    for a in sys.argv[1:]:
        a = a.strip()
        if a == "custom":
            args["custom"] = True
        elif a.startswith("workers="):
            try:
                args["workers"] = max(1, int(a.split("=", 1)[1]))
            except ValueError:
                pass
        elif a.startswith("max_ips="):
            try:
                args["max_ips"] = max(1, int(a.split("=", 1)[1]))
            except ValueError:
                pass
        elif a.startswith("report="):
            args["report"] = a.split("=", 1)[1].strip()
        elif a.startswith("clear_temp="):
            v = a.split("=", 1)[1].strip().lower()
            if v in ("1", "true", "yes", "y"):
                args["clear_temp"] = True
            elif v in ("0", "false", "no", "n"):
                args["clear_temp"] = False
    return args


def ask(prompt: str, default: str = "") -> str:
    raw = input(c(f"  {prompt} [{default}]: ", Fore.CYAN)).strip()
    return raw if raw else default


def interactive() -> Dict[str, Any]:
    title("Custom options")
    info("Press Enter to keep default")
    workers = int(ask("Workers", str(DEFAULT_WORKERS)) or DEFAULT_WORKERS)
    max_raw = ask("Max IPs (0=all)", "0")
    max_ips = int(max_raw) if max_raw.isdigit() and int(max_raw) > 0 else None
    report = ask("Report name", "")
    clear = ask("Clear temp after run? (y/n)", "y").lower() in ("y", "yes", "1", "")
    return {"workers": workers, "max_ips": max_ips, "report": report, "clear_temp": clear}


# ==================== Main ====================

def main() -> None:
    global CFG, TEMP_DIR, RESULTS_DIR, LOG_DIR
    os.chdir(SCRIPT_DIR)
    title("CF Xray IP Benchmark")
    print(c("  Cloudflare IP + Xray config tester", Fore.WHITE))
    print()

    CFG = load_config()

    # Apply paths from config
    TEMP_DIR = SCRIPT_DIR / str(cfg_get("paths", "temp_dir", default="temp"))
    RESULTS_DIR = SCRIPT_DIR / str(cfg_get("paths", "results_dir", default="results"))
    LOG_DIR = SCRIPT_DIR / str(cfg_get("paths", "logs_dir", default="logs"))

    args = parse_args()
    if args["custom"]:
        opts = interactive()
        workers = opts["workers"]
        max_ips = opts["max_ips"]
        report = opts["report"]
        clear_temp = opts["clear_temp"]
    else:
        # CLI overrides config; config is default
        workers = args["workers"] if "workers=" in " ".join(sys.argv[1:]) else int(cfg_get("workers", default=1) or 1)
        # if CLI passed workers= keep args; parse_args always sets workers from DEFAULT or CLI
        # simpler: CLI wins when explicitly passed
        cli_has_workers = any(a.startswith("workers=") for a in sys.argv[1:])
        cli_has_max = any(a.startswith("max_ips=") for a in sys.argv[1:])
        cli_has_report = any(a.startswith("report=") for a in sys.argv[1:])
        cli_has_clear = any(a.startswith("clear_temp=") for a in sys.argv[1:])

        workers = args["workers"] if cli_has_workers else int(cfg_get("workers", default=1) or 1)
        max_ips_cfg = cfg_get("max_ips", default=0) or 0
        max_ips = args["max_ips"] if cli_has_max else (int(max_ips_cfg) if int(max_ips_cfg) > 0 else None)
        report = args["report"] if cli_has_report else str(cfg_get("report_name", default="") or "")
        clear_temp = args["clear_temp"] if cli_has_clear else cfg_get("clear_temp", default=None)

    if cfg_get("display", "write_log_file", default=True):
        init_log(report)
    xray_path = ensure_xray()
    ensure_geodata()
    print(c("-" * 64, Fore.BLUE))

    baseline = measure_baseline()

    cfip_path = SCRIPT_DIR / str(cfg_get("paths", "cfip_file", default="cfip.txt"))
    links_path = SCRIPT_DIR / str(cfg_get("paths", "links_file", default="links.txt"))
    cf_ips = load_cf_ips(cfip_path, max_ips)
    links = load_links(links_path)
    if not cf_ips or not links:
        err("Need valid cfip.txt and links.txt")
        sys.exit(1)

    parsed_list: List[Tuple[str, Dict]] = []
    for link in links:
        p = parse_link(link)
        if p:
            parsed_list.append((p["remark"], p))
            ok(f"Config: {p['remark']}  ({p['address']}:{p['port']})")
        else:
            warn("Skipped invalid link")

    if not parsed_list:
        err("No valid configs")
        sys.exit(1)

    tasks = [(remark, parsed, ip) for remark, parsed in parsed_list for ip in cf_ips]
    info(f"Tests={len(tasks)}  workers={workers}  IPs={len(cf_ips)}  configs={len(parsed_list)}")
    print(c("-" * 64, Fore.BLUE))

    results: List[Dict] = []

    if workers <= 1:
        for i, (remark, parsed, ip) in enumerate(tasks, 1):
            res = run_one(xray_path, parsed, ip, remark)
            results.append(res)
            print_result(res, i, len(tasks))
    else:
        with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as pool:
            futs = {
                pool.submit(run_one, xray_path, parsed, ip, remark): (remark, ip)
                for remark, parsed, ip in tasks
            }
            done = 0
            for fut in concurrent.futures.as_completed(futs):
                done += 1
                try:
                    res = fut.result()
                except Exception as e:
                    remark, ip = futs[fut]
                    res = {
                        "name": f"{remark} | {ip}",
                        "cf_ip": ip,
                        "remark": remark,
                        "status": f"FAIL (worker: {e})",
                        "exit_ip": "?",
                        "loc": "?",
                        "colo": "?",
                        "lat_avg": None,
                        "lat_min": None,
                        "lat_max": None,
                        "jitter": None,
                        "loss": None,
                        "dl_avg": None,
                        "dl_min": None,
                        "dl_max": None,
                        "ul_avg": None,
                        "ul_min": None,
                        "ul_max": None,
                        "web_score": 0,
                        "insta_score": 0,
                        "gaming_score": 0,
                        "overall": 0,
                        "relay": {},
                        "port": 0,
                    }
                results.append(res)
                print_result(res, done, len(tasks))

    title("Final Ranking & Summary")
    info("All scores are out of 10  (0 = worst, 10 = best)")
    print()

    ok_res = sorted(
        [r for r in results if r["status"] == "OK"],
        key=lambda x: x["overall"],
        reverse=True,
    )

    def _cell(val, width, align="<"):
        s = "-" if val is None else str(val)
        if align == ">":
            return f"{s:>{width}}"
        if align == "^":
            return f"{s:^{width}}"
        return f"{s:<{width}}"

    def print_table(title_text, color, rows, headers, widths):
        """rows: list of list of plain strings (same len as headers)."""
        print(c(f"  {title_text}", color + Style.BRIGHT))
        sep = "+" + "+".join("-" * (w + 2) for w in widths) + "+"
        hdr = "|" + "|".join(f" {headers[i]:<{widths[i]}} " for i in range(len(headers))) + "|"
        print(c("  " + sep, Fore.BLUE))
        print(c("  " + hdr, Fore.WHITE + Style.BRIGHT))
        print(c("  " + sep, Fore.BLUE))
        for row in rows:
            line = "|" + "|".join(f" {str(row[i]):<{widths[i]}} " for i in range(len(headers))) + "|"
            print("  " + line)
        print(c("  " + sep, Fore.BLUE))
        print()

    # ---- Baseline table ----
    if baseline.get("ok"):
        print_table(
            "Your internet (direct, no proxy)",
            Fore.WHITE,
            [[
                str(baseline.get("lat_avg") or "-"),
                str(baseline.get("dl_avg") or "-"),
                str(baseline.get("ul_avg") or "-"),
                f"{baseline.get('web', 0)}/10",
                f"{baseline.get('insta', 0)}/10",
                f"{baseline.get('gaming', 0)}/10",
                f"{baseline.get('overall', 0)}/10",
            ]],
            ["Lat ms", "DL Mbps", "UL Mbps", "Web", "Insta", "Game", "Overall"],
            [8, 9, 9, 6, 6, 6, 8],
        )
    else:
        warn("Baseline not available")
        print()

    if not ok_res:
        warn("No successful proxy tests")
    else:
        def top_rows(key: str):
            n = int(cfg_get("ranking", "top_n", default=5) or 5)
            ranked = sorted(ok_res, key=lambda x: x.get(key, 0), reverse=True)[:n]
            rows = []
            for i, r in enumerate(ranked, 1):
                score = r.get(key, 0)
                delta = ""
                if baseline.get("ok") and cfg_get("ranking", "compare_with_baseline", default=True):
                    bmap = {
                        "web_score": "web",
                        "insta_score": "insta",
                        "gaming_score": "gaming",
                        "overall": "overall",
                    }
                    bval = baseline.get(bmap[key], 0) or 0
                    d = round(float(score) - float(bval), 1)
                    delta = f"{d:+.1f}"
                rows.append([
                    str(i),
                    r["cf_ip"],
                    f"{score}/10",
                    str(r.get("lat_avg") or "-"),
                    str(r.get("dl_avg") or "-"),
                    str(r.get("ul_avg") or "-"),
                    delta or "-",
                ])
            return rows

        hdrs = ["#", "CF IP", "Score", "Lat ms", "DL Mbps", "UL Mbps", "vs Direct"]
        wids = [3, 17, 7, 8, 9, 9, 10]

        n = int(cfg_get("ranking", "top_n", default=5) or 5)
        if cfg_get("ranking", "show_web_top", default=True) and cfg_get("scoring", "web", default=True):
            print_table(f"Top {n} — Web browsing", Fore.CYAN, top_rows("web_score"), hdrs, wids)
        if cfg_get("ranking", "show_instagram_top", default=True) and cfg_get("scoring", "instagram", default=True):
            print_table(f"Top {n} — Instagram", Fore.MAGENTA, top_rows("insta_score"), hdrs, wids)
        if cfg_get("ranking", "show_gaming_top", default=True) and cfg_get("scoring", "gaming", default=True):
            print_table(f"Top {n} — Gaming", Fore.GREEN, top_rows("gaming_score"), hdrs, wids)
        if cfg_get("ranking", "show_overall_top", default=True):
            print_table(f"Top {n} — Overall (balanced)", Fore.YELLOW, top_rows("overall"), hdrs, wids)

        best = ok_res[0]
        best_web = max(ok_res, key=lambda x: x["web_score"])
        best_insta = max(ok_res, key=lambda x: x["insta_score"])
        best_game = max(ok_res, key=lambda x: x.get("gaming_score", 0))

        summary_rows = [
            ["Overall", best["cf_ip"], f"{best['overall']}/10"],
            ["Web", best_web["cf_ip"], f"{best_web['web_score']}/10"],
            ["Instagram", best_insta["cf_ip"], f"{best_insta['insta_score']}/10"],
            ["Gaming", best_game["cf_ip"], f"{best_game.get('gaming_score', 0)}/10"],
        ]
        print_table(
            "Best IP by category",
            Fore.WHITE,
            summary_rows,
            ["Category", "Best CF IP", "Score"],
            [12, 17, 8],
        )

        # Estimate line
        if baseline.get("ok") and baseline.get("dl_avg") and best.get("dl_avg"):
            ratio = (best["dl_avg"] / baseline["dl_avg"] * 100) if baseline["dl_avg"] else 0
            print(f"  Best proxy download ≈ {c(f'{ratio:.0f}%', Fore.YELLOW)} of your direct speed")
            gap = best["overall"] - (baseline.get("overall") or 0)
            if gap >= -1:
                ok("  Proxy quality is close to your direct line.")
            elif gap >= -3:
                warn("  Some overhead vs direct line — still usable.")
            else:
                warn("  Large gap vs direct — try more/cleaner CF IPs.")
        print()
        info("Score guide:  9-10 excellent | 7-8 good | 5-6 average | 3-4 weak | 0-2 poor")

    print()
    if cfg_get("excel", "enabled", default=True):
        save_excel(results, report)
    else:
        info("Excel export disabled in config.json")
    if _log_path:
        ok(f"Log: {_log_path}")

    if TEMP_DIR.exists():
        files = [f for f in TEMP_DIR.iterdir() if f.is_file()]
        info(f"Temp: {TEMP_DIR} ({len(files)} files)")
        do_clear = clear_temp
        if do_clear is None:
            try:
                ans = input(c("  Delete temp files now? [y/N]: ", Fore.CYAN)).strip().lower()
                do_clear = ans in ("y", "yes", "1")
            except Exception:
                do_clear = False
        if do_clear:
            n = 0
            for f in files:
                try:
                    f.unlink()
                    n += 1
                except Exception:
                    pass
            ok(f"Deleted {n} temp files")
        else:
            warn(f"Temp kept: {TEMP_DIR}")

    print()
    ok("Done.")


if __name__ == "__main__":
    main()
